import pandas as pd
from openpyxl import load_workbook, Workbook
import requests
import unicodedata
from difflib import SequenceMatcher
import os
import json
import io
import zipfile
import math
import tempfile
from fastapi import FastAPI, UploadFile, File, Form, HTTPException
from fastapi.responses import StreamingResponse, JSONResponse
from fastapi.middleware.cors import CORSMiddleware

app = FastAPI(title="Logistica API")

# Vercel Serverless requirements (CORS and tmp directory)
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

CAMINHO_CACHE_IBGE = os.path.join(tempfile.gettempdir(), 'municipios_ibge_cache.json')
# We expect the models to be uploaded with the source code in the root directory
ARQUIVO_MODELO_REGIAO = os.path.join(os.getcwd(), 'Modelo Região.xlsx')
ARQUIVO_MODELO_ROTA = os.path.join(os.getcwd(), 'Modelo Rota.xlsx')
ARQUIVO_MODELO_TDE = os.path.join(os.getcwd(), 'Modelo TDE.xlsx')

NOME_ABA = 'Base'
COL_CIDADE = 'Destino'
COL_UF = 'UF Destino'
COL_PRAZO = 'Prazo'
COL_IBGE = 'Codigo IBGE'

def normalizar(texto):
    if pd.isna(texto): return ""
    texto = str(texto).strip().upper()
    try:
        texto_normalizado = unicodedata.normalize('NFKD', texto).encode('ASCII', 'ignore').decode('utf-8')
    except:
        return texto 
    texto_normalizado = texto_normalizado.replace("'", " ").replace(".", " ").replace("-", " ")
    while "  " in texto_normalizado:
        texto_normalizado = texto_normalizado.replace("  ", " ")
    return texto_normalizado.strip()

def API_Atualizar_Cache_IBGE():
    try:
        r = requests.get("https://servicodados.ibge.gov.br/api/v1/localidades/municipios", timeout=30, verify=False)
        municipios_api = r.json()
        mapa_final = {}
        for m in municipios_api:
            if not isinstance(m, dict): continue
            nome = normalizar(m.get('nome', ''))
            micro = m.get("microrregiao") or {}
            meso = micro.get("mesorregiao") or {}
            uf_obj = meso.get("UF") or {}
            uf = normalizar(uf_obj.get("sigla", ""))
            if nome and uf: 
                mapa_final[(nome, uf)] = {'nome': nome, 'uf': uf, 'id': m.get('id')}
        lista_dados = list(mapa_final.values())
        with open(CAMINHO_CACHE_IBGE, 'w', encoding='utf-8') as f:
            json.dump(lista_dados, f, ensure_ascii=False, indent=4)
        return True
    except Exception as e:
        print(f"Falha ao processar dados IBGE: {e}")
        return False

@app.get("/api/download_base")
def download_base():
    wb = Workbook()
    ws = wb.active
    ws.title = "Base"
    headers = ["Nome da Região", "Destino", "UF Destino", "Prazo", "Codigo IBGE", "DOMINGO", "SEGUNDA", "TERÇA", "QUARTA", "QUINTA", "SEXTA", "SABADO", "FREQUENCIA"]
    ws.append(headers)
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return StreamingResponse(output, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": "attachment; filename=Base_Template.xlsx"})

@app.get("/api/cidades_ibge")
def carregar_lista_cidades_ibge():
    if not os.path.exists(CAMINHO_CACHE_IBGE):
        API_Atualizar_Cache_IBGE()
    try:
        with open(CAMINHO_CACHE_IBGE, 'r', encoding='utf-8') as f:
            dados = json.load(f)
        lista = [f"{m['nome']} - {m['uf']} ({m['id']})" for m in dados]
        return JSONResponse(content={"cidades": sorted(lista)})
    except Exception:
        return JSONResponse(content={"cidades": []})

@app.post("/api/processar_ibge")
async def api_processar_ibge(file: UploadFile = File(...)):
    if not os.path.exists(CAMINHO_CACHE_IBGE):
        API_Atualizar_Cache_IBGE()
    with open(CAMINHO_CACHE_IBGE, 'r', encoding='utf-8') as f:
        municipios_lista = json.load(f)
        
    db_ibge_por_uf = {}
    mapa_exato = {} 
    for m in municipios_lista:
        nome_norm = normalizar(m['nome'])
        uf_norm = normalizar(m['uf'])
        id_ibge = m['id']
        mapa_exato[(nome_norm, uf_norm)] = id_ibge
        if uf_norm not in db_ibge_por_uf: db_ibge_por_uf[uf_norm] = []
        db_ibge_por_uf[uf_norm].append({'nome_norm': nome_norm, 'id': id_ibge, 'nome_real': m['nome']})
        
    content = await file.read()
    file_io = io.BytesIO(content)
    df = pd.read_excel(file_io, sheet_name=NOME_ABA)
    file_io.seek(0) 
    
    if COL_IBGE not in df.columns: df[COL_IBGE] = ""
    wb = load_workbook(file_io)
    ws = wb[NOME_ABA]
    col_ibge_num = df.columns.get_loc(COL_IBGE) + 1 
    
    count_exato, count_aprox = 0, 0
    for index, row in df.iterrows():
        cidade_excel_raw = str(row[COL_CIDADE])
        uf_excel_raw = str(row[COL_UF])
        cidade_norm = normalizar(cidade_excel_raw)
        uf_norm = normalizar(uf_excel_raw)
        ibge_encontrado = mapa_exato.get((cidade_norm, uf_norm))
        if ibge_encontrado:
            count_exato += 1
        else:
            if uf_norm in db_ibge_por_uf:
                candidatos = db_ibge_por_uf[uf_norm]
                melhor_ratio = 0.0
                melhor_candidato = None
                for item in candidatos:
                    ratio = SequenceMatcher(None, cidade_norm, item['nome_norm']).ratio()
                    if ratio > melhor_ratio:
                        melhor_ratio = ratio
                        melhor_candidato = item
                if melhor_ratio >= 0.80 and melhor_candidato:
                    ibge_encontrado = melhor_candidato['id']
                    count_aprox += 1
        if ibge_encontrado:
            ws.cell(row=index + 2, column=col_ibge_num).value = ibge_encontrado
            
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output, 
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", 
        headers={
            "Content-Disposition": "attachment; filename=Base_IBGE_Preenchida.xlsx",
            "X-Stats": f"Exatos: {count_exato} | Aprox: {count_aprox}"
        }
    )

@app.post("/api/processar_prazos")
async def api_processar_prazos(file_destino: UploadFile = File(...), file_base: UploadFile = File(...)):
    content_destino = await file_destino.read()
    content_base = await file_base.read()
    
    file_base_io = io.BytesIO(content_base)
    file_destino_io = io.BytesIO(content_destino)
    
    df_base = pd.read_excel(file_base_io, sheet_name="Base", dtype=str).fillna("")
    df_base.columns = [str(col).strip().upper() for col in df_base.columns]
    
    nome_coluna_texto_freq = None
    for col in df_base.columns:
        amostra = df_base[col].head(20).astype(str).tolist()
        for val in amostra:
            if "....." in val or "STQQS" in val:
                nome_coluna_texto_freq = col; break
        if nome_coluna_texto_freq: break
        
    mapa_colunas_base = {}
    padroes_busca = {'Seg': 'SEG', 'Ter': 'TER', 'Qua': 'QUA', 'Qui': 'QUI', 'Sex': 'SEX', 'Sáb': 'SAB', 'Dom': 'DOM'}
    for dia_destino, texto_busca in padroes_busca.items():
        for col_real in df_base.columns:
            if texto_busca in col_real:
                mapa_colunas_base[dia_destino] = col_real; break
                
    col_ibge_nome = next((c for c in df_base.columns if 'IBGE' in c), 'CODIGO IBGE')
    df_base = df_base.dropna(subset=[col_ibge_nome])
    df_base['IBGE_LIMPO'] = df_base[col_ibge_nome].apply(lambda x: str(x).split('.')[0].strip())
    df_base = df_base.drop_duplicates(subset=['IBGE_LIMPO'], keep='first')
    dicionario_base = df_base.set_index('IBGE_LIMPO').to_dict('index')
    
    wb = load_workbook(file_destino_io)
    nome_aba_destino = "Prazo (localizações)"
    sheet = wb[nome_aba_destino] if nome_aba_destino in wb.sheetnames else wb.active
    header_row = 4
    col_map_dest = {}
    
    if sheet.max_row >= header_row:
        for cell in sheet[header_row]:
            if cell.value: col_map_dest[str(cell.value).strip()] = cell.column
    else: 
        raise HTTPException(status_code=400, detail=f"A planilha de destino não tem dados na linha {header_row}.")
        
    idx_ibge = col_map_dest.get('Código IBGE da Cidade')
    if not idx_ibge:
        for k, v in col_map_dest.items():
            if "IBGE" in k.upper() and "CIDADE" in k.upper():
                idx_ibge = v; break
    if not idx_ibge: 
        raise HTTPException(status_code=400, detail="Coluna de IBGE não encontrada no destino.")
        
    cidades_atualizadas = 0
    for row_index in range(header_row + 1, sheet.max_row + 1):
        cell_ibge = sheet.cell(row=row_index, column=idx_ibge)
        if not cell_ibge.value: continue
        ibge_chave = str(cell_ibge.value).split('.')[0].strip()
        
        if ibge_chave in dicionario_base:
            dados_linha = dicionario_base[ibge_chave]
            if 'Prazo' in col_map_dest and 'PRAZO' in dados_linha:
                sheet.cell(row=row_index, column=col_map_dest['Prazo']).value = dados_linha['PRAZO']
                
            eh_caso_pontinhos = False
            if nome_coluna_texto_freq:
                texto_freq = str(dados_linha.get(nome_coluna_texto_freq, "")).strip()
                if "......" in texto_freq or (len(texto_freq) > 3 and set(texto_freq) == {'.'}): eh_caso_pontinhos = True
                
            if eh_caso_pontinhos:
                for dia in ['Seg', 'Ter', 'Qua', 'Qui', 'Sex', 'Sáb']:
                    if dia in col_map_dest: sheet.cell(row=row_index, column=col_map_dest[dia]).value = 'VERDADEIRO'
                if 'Dom' in col_map_dest: sheet.cell(row=row_index, column=col_map_dest['Dom']).value = 'FALSO'
            else:
                for dia_curto, nome_coluna_base in mapa_colunas_base.items():
                    if dia_curto in col_map_dest:
                        valor_bruto = str(dados_linha.get(nome_coluna_base, "")).upper().strip()
                        eh_verdadeiro = valor_bruto in ['VERDADEIRO', 'TRUE', 'S', 'SIM', '1', 'X']
                        sheet.cell(row=row_index, column=col_map_dest[dia_curto]).value = 'VERDADEIRO' if eh_verdadeiro else 'FALSO'
            cidades_atualizadas += 1
            
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output, 
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", 
        headers={"Content-Disposition": "attachment; filename=Destino_Prazos.xlsx", "X-Stats": f"Cidades Atualizadas: {cidades_atualizadas}"}
    )

@app.post("/api/processar_regiao")
async def api_processar_regiao(file_base: UploadFile = File(...), cnpj: str = Form(...)):
    if not os.path.exists(ARQUIVO_MODELO_REGIAO):
        raise HTTPException(status_code=500, detail="Modelo Região.xlsx não encontrado no servidor.")
        
    content = await file_base.read()
    file_io = io.BytesIO(content)
    
    df_prazos = pd.read_excel(file_io, sheet_name='Base')
    df_prazos['Nome da Região'] = df_prazos['Nome da Região'].astype(str).str.strip()
    df_prazos = df_prazos[df_prazos['Nome da Região'].notna() & (df_prazos['Nome da Região'].str.upper() != 'NAN') & (df_prazos['Nome da Região'] != '')]
    df_prazos['NomeRegiao'] = df_prazos['Nome da Região'].str.upper()
    
    wb_modelo = load_workbook(ARQUIVO_MODELO_REGIAO)
    ws_regioes = wb_modelo['regioes']
    ws_localizacoes = wb_modelo['localizacoes_atendidas']
    
    for ws in [ws_regioes, ws_localizacoes]:
        for row in ws.iter_rows(min_row=5, max_row=ws.max_row):
            for cell in row: cell.value = None
            
    for i, nome_regiao in enumerate(df_prazos['NomeRegiao'].unique(), start=5):
        ws_regioes[f'B{i}'] = cnpj; ws_regioes[f'C{i}'] = nome_regiao; ws_regioes[f'D{i}'] = "VERDADEIRO"
        
    for i, row in enumerate(df_prazos.iterrows(), start=5):
        ws_localizacoes[f'B{i}'] = row[1]['NomeRegiao']; ws_localizacoes[f'E{i}'] = row[1]['Codigo IBGE']
        
    output = io.BytesIO()
    wb_modelo.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output, 
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", 
        headers={"Content-Disposition": "attachment; filename=Regioes.xlsx"}
    )

@app.post("/api/processar_rotas")
async def api_processar_rotas(
    file_modelo_regioes: UploadFile = File(...),
    tipo_rota: str = Form(...),
    cnpj_rota: str = Form(...),
    nome_transp_rota: str = Form(...),
    desc_rota: str = Form(""),
    tipo_origem: str = Form(...),
    valor_origem: str = Form(...)
):
    if not os.path.exists(ARQUIVO_MODELO_ROTA):
        raise HTTPException(status_code=500, detail="Modelo Rota.xlsx não encontrado no servidor.")
        
    content_reg = await file_modelo_regioes.read()
    file_reg_io = io.BytesIO(content_reg)
    
    wb_modelo_regioes = load_workbook(file_reg_io)
    ws_regioes = wb_modelo_regioes['regioes']
    
    regioes_encontradas = [str(ws_regioes.cell(row=i, column=3).value) for i in range(5, ws_regioes.max_row + 1) if ws_regioes.cell(row=i, column=3).value]
    if not regioes_encontradas: 
        raise HTTPException(status_code=400, detail="Nenhuma região encontrada no modelo de regiões.")
        
    wb_rotas = load_workbook(ARQUIVO_MODELO_ROTA)
    ws_rotas = wb_rotas["Rotas"] if "Rotas" in wb_rotas.sheetnames else wb_rotas.active
    
    next_row = 6
    while ws_rotas.cell(row=next_row, column=1).value is not None: 
        next_row += 1
        
    for regiao_destino in regioes_encontradas:
        ws_rotas.cell(row=next_row, column=1).value = f"{cnpj_rota} - {nome_transp_rota}"
        desc = f"{desc_rota} x {regiao_destino}" if desc_rota else regiao_destino
        ws_rotas.cell(row=next_row, column=2).value = desc
        
        if tipo_origem == "Cidade (IBGE)": 
            ws_rotas.cell(row=next_row, column=3).value = valor_origem
        else: 
            ws_rotas.cell(row=next_row, column=5).value = valor_origem
            
        ws_rotas.cell(row=next_row, column=8).value = regiao_destino
        ws_rotas.cell(row=next_row, column=10).value = "VERDADEIRO"
        ws_rotas.cell(row=next_row, column=11).value = "VERDADEIRO"
        next_row += 1
        
    output = io.BytesIO()
    wb_rotas.save(output)
    output.seek(0)
    
    nome_arq = f"Rotas_{nome_transp_rota.strip()}.xlsx" if nome_transp_rota.strip() else "Rotas_Preenchidas.xlsx"
    return StreamingResponse(
        output, 
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", 
        headers={"Content-Disposition": f"attachment; filename={nome_arq}"}
    )

@app.post("/api/converter_sn")
async def api_converter_sn(file: UploadFile = File(...)):
    content = await file.read()
    wb = load_workbook(io.BytesIO(content))
    ws = wb[NOME_ABA]
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=6, max_col=12):
        for cell in row:
            valor = str(cell.value).strip().upper() if cell.value is not None else ""
            if valor == "S": cell.value = "VERDADEIRO"
            elif valor == "N": cell.value = "FALSO"
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return StreamingResponse(
        output, 
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", 
        headers={"Content-Disposition": "attachment; filename=S_N_Convertido.xlsx"}
    )

@app.post("/api/converter_stqqs")
async def api_converter_stqqs(file: UploadFile = File(...)):
    content = await file.read()
    wb = load_workbook(io.BytesIO(content))
    ws = wb.active
    colunas_destino = [7, 8, 9, 10, 11, 12]
    letras_referencia = ['S', 'T', 'Q', 'Q', 'S', 'S']
    for row in ws.iter_rows(min_row=2):
        texto_raw = str(row[12].value or "").strip().upper()
        for i, col_idx in enumerate(colunas_destino):
            if i < len(texto_raw) and texto_raw[i] == letras_referencia[i]: row[col_idx - 1].value = True
            else: row[col_idx - 1].value = False
        row[5].value = False
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return StreamingResponse(
        output, 
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", 
        headers={"Content-Disposition": "attachment; filename=STQQS_Convertido.xlsx"}
    )

@app.post("/api/restricoes")
async def api_restricoes(
    texto_rest: str = Form(...),
    limite_linhas: int = Form(500),
    categoria: str = Form(...),
    tipo_pessoa: str = Form(...),
    usar_valor: bool = Form(True)
):
    if not os.path.exists(ARQUIVO_MODELO_TDE):
        raise HTTPException(status_code=500, detail="Modelo TDE.xlsx não encontrado no servidor.")
        
    linhas = [l for l in texto_rest.split('\n') if l.strip()]
    grupos_por_valor = {}

    for l in linhas:
        partes = l.replace('\t', ' ').strip().split()
        if len(partes) < 2: continue 
        
        cnpj_raw = partes[0].strip().replace('.', '').replace('/', '').replace('-', '')
        possivel_valor = partes[-1].replace("R$", "").replace(".", "").replace(",", ".")
        valor_final = "0"
        razao_partes = partes[1:]

        try:
            float(possivel_valor)
            valor_final = possivel_valor
            razao_partes = partes[1:-1]
        except:
            valor_final = "0" 

        razao = " ".join(razao_partes).upper()
        chave_grupo = valor_final if usar_valor else "GERAL"

        if chave_grupo not in grupos_por_valor:
            grupos_por_valor[chave_grupo] = []
        grupos_por_valor[chave_grupo].append((cnpj_raw, razao, valor_final))

    with open(ARQUIVO_MODELO_TDE, "rb") as f:
        template_bytes = io.BytesIO(f.read())
        
    zip_buffer = io.BytesIO()
    with zipfile.ZipFile(zip_buffer, "a", zipfile.ZIP_DEFLATED, False) as zip_file:
        for valor_chave, dados_lista in grupos_por_valor.items():
            total_registros = len(dados_lista)
            num_arquivos = math.ceil(total_registros / limite_linhas)
            
            for i in range(num_arquivos):
                inicio = i * limite_linhas
                fim = min((i + 1) * limite_linhas, total_registros)
                lote = dados_lista[inicio:fim]
                
                template_bytes.seek(0)
                wb = load_workbook(template_bytes)
                ws = wb["Pessoa"] if "Pessoa" in wb.sheetnames else wb.active

                row_idx = 5
                for cnpj, razao, v_lin in lote:
                    ws.cell(row=row_idx, column=1).value = str(cnpj)
                    ws.cell(row=row_idx, column=2).value = str(tipo_pessoa) 
                    ws.cell(row=row_idx, column=4).value = str(razao)
                    row_idx += 1

                prefixo = f"{categoria} " if categoria != "Outros" else ""
                faixa_linhas = f"{inicio} a {fim} Linhas"
                sufixo_valor = f" R${valor_chave}" if (usar_valor and valor_chave != "0") else ""
                
                nome_arquivo = f"{prefixo}{faixa_linhas}{sufixo_valor}.xlsx".strip()
                
                excel_buffer = io.BytesIO()
                wb.save(excel_buffer)
                excel_buffer.seek(0)
                zip_file.writestr(nome_arquivo, excel_buffer.read())

    zip_buffer.seek(0)
    return StreamingResponse(
        zip_buffer,
        media_type="application/zip",
        headers={"Content-Disposition": f"attachment; filename=Restricoes_{categoria}.zip"}
    )
